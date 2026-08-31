# ==============================================================================
# main.py — Orchestrator (full pipeline)
# ==============================================================================
#
# PHASE A (once per run):
#   1. Selenium → ICISnet scraping → df_final
#   2. Pywinauto → SAP export → ΕΙΣΑΓΩΓΕΣ_database.xlsx
#   3. Queries → FULL_RESULTS.xlsx (opened / done / undone)
#
# PAUSE — Editable approval popup:
#   Shows: MRN, PROT, PDF, PDF.1, STATUS
#   Checkbox per row | Double-click → inline edit | OK → start Phase B
#
# PHASE B (for each selected MRN):
#   4. Selenium → ICISnet → ID29 → download XML → browser quit
#   5. Parse XML → df_result + filters
#   6. PDF rename (pdf1 → pdf)
#   7. SAP entry loop (login only on first item per MRN)
#   8. sap_attach + wait for confirmation
#   9. SAP closes → PDF move → archive folder
#  10. Status → DONE in FULL_RESULTS.xlsx
#
# ==============================================================================

import os
import sys
import io
import ssl
import time
import shutil
import base64
import warnings
import subprocess
import traceback
import ctypes
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from io import StringIO
from datetime import datetime
from time import perf_counter
from dateutil.relativedelta import relativedelta

import pyautogui
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from pywinauto import Application
from pywinauto.keyboard import send_keys

# Encoding for Power Automate compatibility
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

# ==============================================================================
# SETTINGS
# ==============================================================================

os.environ["WDM_SSL_VERIFY"] = "0"
ssl._create_default_https_context = ssl._create_unverified_context
warnings.filterwarnings("ignore")

ICISNET_USER  = "YOUR_ICISNET_USERNAME"
ICISNET_PASS  = "YOUR_ICISNET_PASSWORD"
SAP_USER      = "YOUR_SAP_USERNAME"
SAP_PASS      = "YOUR_SAP_PASSWORD"

_BASE         = Path(r"C:\Users\YOUR_USERNAME\OneDrive - YOUR_COMPANY")
DESKTOP       = _BASE / "Desktop"
DOCUMENTS     = _BASE / "Documents"
SAP_FILE_A    = DESKTOP / "LIST_N.sap"
SAP_FILE_B    = DESKTOP / "ZELVMM_IMP_1.sap"
GERAKARHS     = DESKTOP / "MRN_REFERENCE" / "MRN.xlsx"
DATABASE_XLSX = DESKTOP / "ΕΙΣΑΓΩΓΕΣ_database.xlsx"
SAP_GUI_DIR   = DOCUMENTS / "SAP" / "SAP GUI"
PDF_SAVE_PATH = DESKTOP / "Αναζήτηση _ Αποτελέσματα Αναζήτησης.pdf"
OUTPUT_EXCEL  = DESKTOP / "python" / "FULL_RESULTS.xlsx"
SAVE_FOLDER   = DESKTOP / "python" / "xml_temp"
XML_PATH      = SAVE_FOLDER / "current.xml"
ATLAS_BASE    = Path(r"\\YOUR_SERVER\YOUR_SHARE\ΤΕΛΩΝΕΙΑ\ΗΛΕΚΤΡΟΝΙΚΟ ΑΡΧΕΙΟ ΔΙΑΣΑΦΗΣΕΩΝ ΕΙΣΑΓΩΓΩΝ")

KATH_DIR = {
    3:  "ΑΠΑΛΛΑΓΗ ΦΠΑ",
    2:  "ΕΝΕΡΓΗΤΙΚΗ",
    5:  "ΕΛΕΥΘΕΡΑ",
    12: "ΕΝΕΡΓΗΤΙΚΗ - INF",
}

_now      = datetime.now()
DATE_TO   = _now.strftime("%d.%m.%Y")
DATE_FROM = (_now.replace(day=1) - relativedelta(months=5)).strftime("01.%m.%Y")

# ⚠️  Update this each time the processing period changes
LIMIT_DATE = pd.Timestamp(2025, 8, 1).date()

SHOW_COLS     = ["MRN", "PROT", "PDF", "PDF.1", "ΚΑΤΑΣΤΑΣΗ"]
EDIT_COLS     = {"PROT", "PDF", "PDF.1"}
ALLOWED_DASMOS = ("76", "72", "81", "2804690", "2710112")

# ==============================================================================
# HELPERS
# ==============================================================================

def make_chrome(download_folder: Path = None):
    """Build Chrome driver. If download_folder is provided → auto-download mode."""
    opts = Options()
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--start-maximized")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    if download_folder:
        opts.add_experimental_option("prefs", {
            "download.default_directory": str(download_folder),
            "download.prompt_for_download": False,
            "safebrowsing.enabled": True,
        })
    else:
        opts.add_argument("--kiosk-printing")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


def close_sap():
    for proc in ["saplogon.exe", "saplgpad.exe", "sapgui.exe", "nwbc.exe"]:
        subprocess.call(["taskkill", "/F", "/T", "/IM", proc],
                        stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)


def setup_keyboard():
    hwnd = ctypes.windll.user32.GetForegroundWindow()
    ctypes.windll.user32.PostMessageW(hwnd, 0x0050, 0, 0x0408)
    if ctypes.windll.user32.GetKeyState(0x14) & 1:
        pyautogui.press("capslock")


def safe_select(wait, element_id, text):
    for _ in range(5):
        try:
            Select(wait.until(EC.element_to_be_clickable((By.ID, element_id))))\
                .select_by_visible_text(text)
            return
        except:
            time.sleep(1)
    raise Exception(f"safe_select failed: {element_id}={text}")


def popup_input(title: str, prompt: str, default: str = "") -> str:
    from tkinter import simpledialog
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    result = simpledialog.askstring(title, prompt, initialvalue=default, parent=root)
    root.destroy()
    return result.strip() if result else ""

def show_info(msg: str):
    """Show informational banner with OK button."""
    root = tk.Tk()
    root.title("Ενημέρωση")
    root.resizable(False, False)
    root.configure(bg="#F5F4F0")
    root.attributes("-topmost", True)
    w, h = 420, 150
    sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
    root.lift(); root.focus_force()

    tk.Label(root, text=msg,
        bg="#F5F4F0", fg="#1A1A1A", font=("Consolas", 10),
        wraplength=380, justify="center"
    ).pack(pady=(25, 15))

    tk.Button(root, text="OK", font=("Consolas", 10, "bold"),
        bg="#1A1A1A", fg="#FFFFFF", relief="flat", padx=20, pady=6,
        cursor="hand2", command=root.destroy
    ).pack()

    root.mainloop()

# ==============================================================================
# PHASE A.1 — ICISnet scraping
# ==============================================================================

def phase_a_icisnet() -> pd.DataFrame:
    for attempt in range(1, 4):
        print(f"  ICISnet — attempt {attempt}/3...")
        driver = None
        try:
            driver = make_chrome()
            wait   = WebDriverWait(driver, 45)

            driver.get("https://www1.gsis.gr/icisnet/itrader/common/home.jsf")
            wait.until(EC.element_to_be_clickable((By.NAME, "username"))).send_keys(ICISNET_USER)
            driver.find_element(By.NAME, "password").send_keys(ICISNET_PASS)
            driver.find_element(By.NAME, "btn_login").click()
            wait.until(EC.url_contains("icisnet"))

            wait.until(EC.element_to_be_clickable((By.ID, "iconcontentForm:mainMenu_ics"))).click()
            wait.until(EC.element_to_be_clickable((By.ID, "iconcontentForm:ics_import_declaration"))).click()
            wait.until(EC.element_to_be_clickable((By.ID, "iconcontentForm:menu_ics_import_declaration_search"))).click()

            df_el = wait.until(EC.element_to_be_clickable(
                (By.ID, "contentForm:submission_date_fromInputDate")))
            driver.execute_script(
                "arguments[0].removeAttribute('readonly'); arguments[0].value=arguments[1];",
                df_el, DATE_FROM.replace(".", "-"))
            Select(driver.find_element(By.ID, "contentForm:search_scope")).select_by_value("Trader")

            for sa in range(1, 11):
                try:
                    driver.find_element(By.XPATH, "//input[@value='Αναζήτηση']").click()
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.ID, "contentForm:printResultsReport")))
                    break
                except:
                    if sa == 10: raise
                    time.sleep(1)

            driver.execute_script(
                "jsfcljs(document.getElementById('contentForm'),{"
                "'contentForm:printResultsReport':'contentForm:printResultsReport',"
                "'showLrn':'true','dispatch':'','movementReferenceLabel':'MRN',"
                "'noLRNcol':'false'},'new');")
            wait.until(lambda d: len(d.window_handles) > 1)
            driver.switch_to.window(driver.window_handles[-1])
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Αποτελέσματα Αναζήτησης')]")))

            try:
                pdf_data = driver.execute_cdp_cmd("Page.printToPDF", {"printBackground": True})
                PDF_SAVE_PATH.write_bytes(base64.b64decode(pdf_data["data"]))
                print("    PDF saved")
            except Exception as e:
                print(f"    PDF save failed: {e}")

            try:
                df_raw = pd.read_html(StringIO(driver.page_source))[0]
            except ValueError:
                driver.quit()
                print("    ICISnet: no import declarations found")
                return pd.DataFrame(columns=["MRN","ΤΥΠΟΣ","ΚΑΤΑΣΤΑΣΗ","LRN","ΗΜ_ΥΠΟΒ","ΗΜ_ΕΝΗΜ","PDF"])

            driver.quit(); driver = None

            # Clean and filter results
            df = df_raw.copy()
            df.columns = [str(c).strip() for c in df.columns]
            repl = [
                ("YOUR_LRN_PREFIX_ALT/", "YOUR_LRN_PREFIX/"), ("ELVELV", "ELV"),
                ("YOUR_LRN_PREFIX_ALT /", "ELV800924063/25 /"),
                ("YOUR_VAT_PREFIX_WRONG", "YOUR_VAT_PREFIX_CORRECT"),
                ("YOUR_CB_PREFIX_SHORT", "YOUR_CB_PREFIX_LONG"), ("YOUR_CB_ALT", "YOUR_CB_ALT_CORRECT"),
            ]
            df["LRN"] = df["LRN"].astype(str)
            for old, new in repl:
                df["LRN"] = df["LRN"].str.replace(old, new, regex=False)
            df["Τύπος Δήλωσης"] = df["Τύπος Δήλωσης"].astype(str).str.replace("-","",regex=False)
            df["Ημ/νία Υποβολής"] = pd.to_datetime(
                df["Ημ/νία Υποβολής"], dayfirst=True, errors="coerce").dt.date
            df["Ημ/νία Ενημέρωσης Κατάστασης"] = pd.to_datetime(
                df["Ημ/νία Ενημέρωσης Κατάστασης"], dayfirst=True, errors="coerce").dt.date
            df = df[~df["LRN"].str.contains("EXCLUDED_LRN_PREFIX", na=False)]
            df = df[
                df["LRN"].str.contains(r"ELV|ΕLV", na=False) |
                df["LRN"].str.contains("YOUR_SPECIFIC_LRN", na=False) |
                df["MRN"].isin(["EXAMPLE_MRN_1","EXAMPLE_MRN_2"])
            ]
            df = df[
                (df["Ημ/νία Υποβολής"] >= LIMIT_DATE) &
                (df["Ημ/νία Ενημέρωσης Κατάστασης"] >= LIMIT_DATE)
            ]
            df["Κατάσταση_Temp"] = df["Κατάσταση"].astype(str).str.replace(
                r"Εισαγωγή.*","ID29", regex=True)
            cond = (
                df["Κατάσταση_Temp"].str.startswith("ID29", na=False) |
                (df["Κατάσταση"] == "Τακτοποιημένο") |
                df["Κατάσταση"].str.contains("Αποδεκτή", na=False) |
                (df["MRN"] == "EXAMPLE_MRN_3")
            )
            df = df[cond]
            df["Κατάσταση"] = df["Κατάσταση_Temp"]
            df["PDF"] = df["MRN"].astype(str) + " " + df["Τύπος Δήλωσης"].astype(str)
            df = df.rename(columns={
                "Τύπος Δήλωσης":"ΤΥΠΟΣ","Κατάσταση":"ΚΑΤΑΣΤΑΣΗ",
                "Ημ/νία Υποβολής":"ΗΜ_ΥΠΟΒ","Ημ/νία Ενημέρωσης Κατάστασης":"ΗΜ_ΕΝΗΜ"})
            df_final = df[["MRN","ΤΥΠΟΣ","ΚΑΤΑΣΤΑΣΗ","LRN","ΗΜ_ΥΠΟΒ","ΗΜ_ΕΝΗΜ","PDF"]]\
                .sort_values(["ΗΜ_ΥΠΟΒ","MRN"], ascending=[False,True]).reset_index(drop=True)
            print(f"    ICISnet OK — {len(df_final)} records")
            return df_final

        except Exception as e:
            if driver:
                try: driver.quit()
                except: pass
            print(f"    Failed: {e}")
            time.sleep(3)

    raise RuntimeError("ICISnet scraping failed after 3 attempts.")


# ==============================================================================
# PHASE A.2 — SAP export
# ==============================================================================

def phase_a_sap_export():
    for attempt in range(1, 4):
        print(f"  SAP Export — attempt {attempt}/3...")
        try:
            close_sap(); time.sleep(3)
            subprocess.Popen(f'start "" "{SAP_FILE_A}"', shell=True)

            app = Application(backend="uia").connect(title="LIST_N", timeout=60)
            win = app.window(title="LIST_N")
            win.child_window(auto_id="1004", control_type="Edit")\
               .wait("exists enabled visible", timeout=30).set_text(SAP_USER)
            win.child_window(auto_id="1005", control_type="Edit").set_text(SAP_PASS)
            win.child_window(title="Εισ.σε Σύστ.", control_type="Button").click_input()

            app2 = Application(backend="uia").connect(title="ΔΙΑΣΑΦΗΣΗ ΕΙΣΑΓΩΓΩΝ", timeout=60)
            win2 = app2.window(title="ΔΙΑΣΑΦΗΣΗ ΕΙΣΑΓΩΓΩΝ")
            win2.wait("ready", timeout=30).set_focus(); win2.maximize()
            send_keys("{TAB 6}" + DATE_FROM + "{TAB}" + DATE_TO + "{TAB 5}1100")
            win2.child_window(title="Εκτέλεση", control_type="Button").click_input()
            time.sleep(10)

            win2.click_input(coords=(51, 21))
            send_keys("{DOWN 3}{RIGHT}{DOWN}{ENTER}"); time.sleep(2)
            send_keys("{ENTER}"); time.sleep(3)

            app_s = Application(backend="win32").connect(title="Save As", timeout=15)
            dlg = app_s.window(title="Save As"); dlg.set_focus()
            try:
                dlg.child_window(class_name="ToolbarWindow32", found_index=1)\
                   .button("Desktop").click_input()
            except: pass
            dlg.child_window(class_name="Edit").set_edit_text("ΕΙΣΑΓΩΓΕΣ_database")
            dlg.child_window(title="&Save", class_name="Button").click_input()
            try:
                c = Application(backend="win32").connect(title="Confirm Save As", timeout=3)
                d = c.window(title="Confirm Save As")
                if d.exists(): d.child_window(title="&Yes", class_name="Button").click_input()
            except: pass

            time.sleep(5)
            subprocess.call(["taskkill","/F","/IM","excel.exe"],
                            stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            time.sleep(2); close_sap()
            print("    SAP Export OK"); return

        except Exception as e:
            print(f"    Failed: {e}"); time.sleep(5)

    raise RuntimeError("SAP Export failed after 3 attempts.")

# ==============================================================================
# PHASE A.3 — Queries → FULL_RESULTS.xlsx
# ==============================================================================

def phase_a_queries(df_final: pd.DataFrame) -> pd.DataFrame:
    print("  Building queries...")
    df_ger = pd.read_excel(GERAKARHS, sheet_name=0)
    df_db  = pd.read_excel(DATABASE_XLSX, sheet_name=0)

    df_ger["MRN"]       = df_ger["MRN"].astype(str)
    df_ger["PROT"]      = df_ger["PROT"].astype(str)
    df_ger["TYPE"]      = df_ger["TYPE"].fillna("").astype(str)
    df_ger["ABC"]       = df_ger["ABC"].fillna("").astype(str)
    df_ger["KATH"]      = df_ger["KATH"].fillna("").astype(str).str.replace(" ","",regex=False)
    df_ger["TYPE_FULL"] = df_ger["TYPE"] + df_ger["ABC"]
    df_ger["PDF"]       = (df_ger["MRN"] + " " + df_ger["TYPE_FULL"]).str.strip()
    df_ger = df_ger.drop_duplicates(subset=["MRN","PROT","TYPE_FULL","PDF"])

    df_db["Καθεστώς Εισαγωγής"] = pd.to_numeric(df_db["Καθεστώς Εισαγωγής"], errors="coerce")
    df_sap_exp = (
        df_db[df_db["Καθεστώς Εισαγωγής"].notna() &
              ~df_db["Καθεστώς Εισαγωγής"].isin([7,4])]
        .rename(columns={"Διασάφηση Εισαγωγής":"MRN"})[["MRN","Ημ/νία δημιουρ."]]
    )
    pdf_files = (
        [f.replace(".pdf","") for f in os.listdir(SAP_GUI_DIR) if "GRIM" in f]
        if SAP_GUI_DIR.exists() else []
    )
    df_gui = pd.DataFrame({"PDF NAME": pdf_files})
    df_gui["PDF NAME"] = df_gui["PDF NAME"].astype(str)

    if df_gui.empty:
        show_info("⚠️  No PDF files found in SAP GUI folder.\nNo declarations to process.")
        df_gui = pd.DataFrame({"PDF NAME": pd.Series([], dtype=str)})

    df_full = df_db.copy()
    df_full["Τελωνείο Εισόδου"] = df_full["Τελωνείο Εισόδου"].astype(str)
    df_full = df_full[
        (df_full["Τελωνείο Εισόδου"] != "0") &
        df_full["Καθεστώς Εισαγωγής"].notna() &
        (df_full["Καθεστώς Εισαγωγής"] != 8)
    ][["Διασάφηση Εισαγωγής","Διασάφηση Εισόδου"]]
    df_full = df_full[
        ~df_full["Διασάφηση Εισαγωγής"].astype(str).str.startswith("23GRIM",na=False)]

    # opened
    o = df_final.copy()
    o = o.merge(df_sap_exp[["MRN"]], on="MRN", how="left", indicator=True)
    o = o[o["_merge"]=="left_only"].drop(columns=["_merge"])
    o = o[~o["ΤΥΠΟΣ"].astype(str).str.endswith(("X","Y","Χ","Υ"),na=False)]
    o = o[~o["MRN"].astype(str).str.contains("GRIM0304",na=False)]
    o = o[~o["MRN"].isin(["EXCLUDED_MRN_4","EXCLUDED_MRN_5"])]
    o = o.merge(df_ger[["MRN","PROT","KATH","PDF"]], on="MRN", how="left", suffixes=("",".1"))
    o = o[o["PDF.1"].notna()]
    o = o[o["KATH"].astype(str) != "7100"]
    o = o.merge(df_full[["Διασάφηση Εισόδου"]], left_on="MRN",
                right_on="Διασάφηση Εισόδου", how="left")
    o = o[o["Διασάφηση Εισόδου"].isna()].drop(columns=["Διασάφηση Εισόδου"])
    o = o.merge(df_gui[["PDF NAME"]], left_on="PDF", right_on="PDF NAME", how="left")
    o["OK"] = o.apply(
        lambda x: "OK" if (pd.notna(x["PDF.1"]) and pd.notna(x["PDF NAME"])
                           and x["PDF"]==x["PDF.1"]==x["PDF NAME"])
        else (None if pd.isna(x["PDF.1"]) else "CHANGE"), axis=1)
    df_opened = o[["MRN","ΤΥΠΟΣ","ΚΑΤΑΣΤΑΣΗ","LRN","ΗΜ_ΥΠΟΒ","ΗΜ_ΕΝΗΜ","PROT","PDF","PDF.1","OK"]]\
        .sort_values(["OK","ΗΜ_ΥΠΟΒ"], ascending=[False,False])

    # done
    d = df_final.copy()
    d = d.merge(df_sap_exp[["MRN"]], on="MRN", how="left", indicator=True)
    d = d.merge(df_full, left_on="MRN", right_on="Διασάφηση Εισόδου", how="left")
    d["NEW_MRN"] = d.apply(
        lambda x: x["MRN"] if x["_merge"]=="both" else x["Διασάφηση Εισαγωγής"], axis=1)
    df_done = d[d["NEW_MRN"].notna()].drop(
        columns=["_merge","Διασάφηση Εισαγωγής","Διασάφηση Εισόδου","NEW_MRN"], errors="ignore")

    # undone
    u = df_final.copy()
    u = u[~u["ΤΥΠΟΣ"].astype(str).str.endswith(("X","Y","Χ","Υ"),na=False)]
    u = u.merge(df_sap_exp[["MRN"]], on="MRN", how="left", indicator=True)
    u = u.merge(df_full, left_on="MRN", right_on="Διασάφηση Εισόδου", how="left")
    u["NEW_MRN"] = u.apply(
        lambda x: x["MRN"] if x["_merge"]=="both" else x["Διασάφηση Εισαγωγής"], axis=1)
    df_undone = u[u["NEW_MRN"].isna()].drop(
        columns=["_merge","Διασάφηση Εισαγωγής","Διασάφηση Εισόδου","NEW_MRN"], errors="ignore")

    # Export FULL_RESULTS.xlsx
    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        for sname, df in {"opened":df_opened,"done":df_done,"undone":df_undone}.items():
            df.to_excel(writer, sheet_name=sname, index=False)
            ws = writer.sheets[sname]; rc, cc = df.shape
            for ci, col in enumerate(ws.columns, 1):
                mx = 0; cl = get_column_letter(ci)
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    try:
                        if cell.value: mx = max(mx, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[cl].width = mx + 3
            if rc > 0:
                ref = f"A1:{get_column_letter(cc)}{rc+1}"
                tab = Table(displayName=sname, ref=ref)
                tab.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium7", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(tab)

    print(f"    FULL_RESULTS.xlsx — opened={len(df_opened)} | done={len(df_done)} | undone={len(df_undone)}")
    return df_opened

# ==============================================================================
# APPROVAL POPUP — Editable table
# ==============================================================================

# Columns visible in the popup
SHOW_COLS = ["MRN", "PROT", "PDF", "PDF.1", "ΚΑΤΑΣΤΑΣΗ",
             "ΚΡΑΜΑ_1", "ΒΑΡΟΣ_1", "ΚΡΑΜΑ_2", "ΒΑΡΟΣ_2", "ΚΡΑΜΑ_3", "ΒΑΡΟΣ_3"]

# Editable columns
EDIT_COLS = {"PROT", "PDF", "PDF.1",
             "ΚΡΑΜΑ_1", "ΒΑΡΟΣ_1", "ΚΡΑΜΑ_2", "ΒΑΡΟΣ_2", "ΚΡΑΜΑ_3", "ΒΑΡΟΣ_3"}


class ApprovalPopup:
    """
    Πρώτο popup — pending MRNs από το FULL_RESULTS.
    Για 0832 + ΔΑΣΜ_ΚΛ 76012080/76012030 συμπληρώνεις ΚΡΑΜΑ/ΒΑΡΟΣ.
    Double-click σε editable στήλη → inline edit.
    Επιστρέφει list[dict] ή None αν ακυρωθεί.
    """

    def __init__(self, rows: list):
        self.rows         = [r.copy() for r in rows]
        self.result       = None
        self._edit_widget = None
        self.root = tk.Tk()
        self.root.title("Έλεγχος Εισαγωγών — Επιλογή & Επεξεργασία")
        self.root.resizable(True, True)
        w, h = 1400, 520
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.root.configure(bg="#F5F4F0")
        self._build()
        self.root.lift()  # ← νέο
        self.root.focus_force()  # ← νέο
        self.root.attributes("-topmost", True)  # ← νέο

    def _build(self):
        r = self.root

        # Header
        hdr = tk.Frame(r, bg="#1A1A1A", height=52)
        hdr.pack(fill="x")
        tk.Label(hdr,
            text="  Pending Εισαγωγές — έλεγξε, τροποποίησε αν χρειαστεί, πάτα OK",
            bg="#1A1A1A", fg="#FFFFFF", font=("Consolas", 11), anchor="w"
        ).pack(side="left", padx=8, pady=14)
        tk.Label(hdr, text=f"{len(self.rows)} MRNs",
            bg="#1A1A1A", fg="#8A8A82", font=("Consolas", 10)
        ).pack(side="right", padx=16)

        # Hint
        hint = tk.Frame(r, bg="#F5F4F0")
        hint.pack(fill="x", padx=14, pady=(8, 2))
        tk.Label(hint,
            text="  Double-click για επεξεργασία   |   "
                 "Για 0832: συμπλήρωσε ΚΡΑΜΑ/ΒΑΡΟΣ (αν 1 κράμα αφησε ΒΑΡΟΣ κενό)",
            bg="#F5F4F0", fg="#6B6B65", font=("Consolas", 9), anchor="w"
        ).pack(side="left")

        # Treeview
        frm = tk.Frame(r, bg="#F5F4F0")
        frm.pack(fill="both", expand=True, padx=14, pady=(4, 6))

        style = ttk.Style(); style.theme_use("clam")
        style.configure("T.Treeview",
            background="#FFFFFF", foreground="#1A1A1A", rowheight=28,
            fieldbackground="#FFFFFF", font=("Consolas", 10), borderwidth=0)
        style.configure("T.Treeview.Heading",
            background="#E8E6DF", foreground="#3A3A36",
            font=("Consolas", 10, "bold"), relief="flat")
        style.map("T.Treeview", background=[("selected", "#D4E8FF")])

        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        all_cols = ["_check"] + SHOW_COLS
        self.tree = ttk.Treeview(frm, columns=all_cols, show="headings",
            style="T.Treeview", yscrollcommand=vsb.set, xscrollcommand=hsb.set,
            selectmode="browse")
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        col_w = {
            "_check": 40, "MRN": 200, "PROT": 70, "PDF": 180, "PDF.1": 180,
            "ΚΑΤΑΣΤΑΣΗ": 90, "ΚΡΑΜΑ_1": 70, "ΒΑΡΟΣ_1": 70,
            "ΚΡΑΜΑ_2": 70, "ΒΑΡΟΣ_2": 70, "ΚΡΑΜΑ_3": 70, "ΒΑΡΟΣ_3": 70,
        }
        col_l = {
            "_check": "ok", "MRN": "MRN", "PROT": "PROT",
            "PDF": "PDF (target)", "PDF.1": "PDF.1 (source)", "ΚΑΤΑΣΤΑΣΗ": "ΚΑΤΑΣΤΑΣΗ",
            "ΚΡΑΜΑ_1": "ΚΡ_1", "ΒΑΡΟΣ_1": "ΒΑΡ_1",
            "ΚΡΑΜΑ_2": "ΚΡ_2", "ΒΑΡΟΣ_2": "ΒΑΡ_2",
            "ΚΡΑΜΑ_3": "ΚΡ_3", "ΒΑΡΟΣ_3": "ΒΑΡ_3",
        }
        for c in all_cols:
            self.tree.heading(c, text=col_l[c])
            self.tree.column(c, width=col_w[c],
                anchor="center" if c == "_check" else "w",
                stretch=(c in ("MRN", "PDF", "PDF.1")))

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.grid_rowconfigure(0, weight=1); frm.grid_columnconfigure(0, weight=1)

        self._cvars = []; self._iids = []
        for i, row in enumerate(self.rows):
            var = tk.BooleanVar(value=True); self._cvars.append(var)
            vals = ["[X]"] + [str(row.get(c, "")) for c in SHOW_COLS]
            iid = self.tree.insert("", "end", iid=str(i), values=vals,
                                   tags=("odd" if i % 2 else "even",))
            self._iids.append(iid)
        self.tree.tag_configure("odd",  background="#FAFAF8")
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.bind("<Button-1>",        self._on_click)
        self.tree.bind("<Double-Button-1>", self._on_dbl)

        # Footer
        foot = tk.Frame(r, bg="#F5F4F0")
        foot.pack(fill="x", padx=14, pady=(0, 12))
        tk.Button(foot, text="Επιλογή Όλων", font=("Consolas", 9),
            bg="#E8E6DF", fg="#1A1A1A", relief="flat", padx=10, pady=5,
            cursor="hand2", command=self._all
        ).pack(side="left", padx=(0, 6))
        tk.Button(foot, text="Αποεπιλογή Όλων", font=("Consolas", 9),
            bg="#E8E6DF", fg="#1A1A1A", relief="flat", padx=10, pady=5,
            cursor="hand2", command=self._none
        ).pack(side="left")
        self._cnt = tk.Label(foot, text="", bg="#F5F4F0", fg="#6B6B65", font=("Consolas", 9))
        self._cnt.pack(side="left", padx=14); self._upd()
        tk.Button(foot, text="Ακύρωση", font=("Consolas", 10),
            bg="#E8E6DF", fg="#6B6B65", relief="flat", padx=14, pady=7,
            cursor="hand2", command=self._cancel
        ).pack(side="right", padx=(6, 0))
        tk.Button(foot, text="OK — Εκκίνηση", font=("Consolas", 10, "bold"),
            bg="#1A1A1A", fg="#FFFFFF", relief="flat", padx=14, pady=7,
            cursor="hand2", activebackground="#333", activeforeground="#FFF",
            command=self._ok
        ).pack(side="right")

    def _on_click(self, event):
        if (self.tree.identify_column(event.x) == "#1" and
                self.tree.identify_region(event.x, event.y) == "cell"):
            iid = self.tree.identify_row(event.y)
            if not iid: return
            self._commit()
            idx = self._iids.index(iid)
            nv  = not self._cvars[idx].get(); self._cvars[idx].set(nv)
            v   = list(self.tree.item(iid, "values")); v[0] = "[X]" if nv else "[ ]"
            self.tree.item(iid, values=v); self._upd()

    def _on_dbl(self, event):
        col_id = self.tree.identify_column(event.x)
        iid    = self.tree.identify_row(event.y)
        if not iid: return
        ci = int(col_id.replace("#", "")) - 1
        ac = ["_check"] + SHOW_COLS
        if ci < 0 or ci >= len(ac): return
        cn = ac[ci]
        if cn not in EDIT_COLS: return
        self._commit()
        bbox = self.tree.bbox(iid, column=col_id)
        if not bbox: return
        x, y, w, h = bbox
        vals = list(self.tree.item(iid, "values"))
        e = tk.Entry(self.tree, font=("Consolas", 10), bg="#FFFDE7",
            fg="#1A1A1A", relief="solid", bd=1, insertbackground="#1A1A1A")
        e.place(x=x, y=y, width=w, height=h)
        e.insert(0, vals[ci]); e.select_range(0, "end"); e.focus_set()
        def commit(ev=None):
            vals[ci] = e.get().strip()
            self.tree.item(iid, values=vals)
            self.rows[self._iids.index(iid)][cn] = vals[ci]
            e.destroy(); self._edit_widget = None
        e.bind("<Return>", commit); e.bind("<Tab>", commit)
        e.bind("<Escape>", lambda ev: e.destroy()); e.bind("<FocusOut>", commit)
        self._edit_widget = e

    def _commit(self):
        if self._edit_widget:
            try: self._edit_widget.event_generate("<FocusOut>")
            except: pass

    def _upd(self):
        n = sum(v.get() for v in self._cvars)
        self._cnt.config(text=f"{n} / {len(self.rows)} επιλεγμένα")

    def _all(self):
        for i, v in enumerate(self._cvars):
            v.set(True)
            vs = list(self.tree.item(self._iids[i], "values")); vs[0] = "[X]"
            self.tree.item(self._iids[i], values=vs)
        self._upd()

    def _none(self):
        for i, v in enumerate(self._cvars):
            v.set(False)
            vs = list(self.tree.item(self._iids[i], "values")); vs[0] = "[ ]"
            self.tree.item(self._iids[i], values=vs)
        self._upd()

    def _ok(self):
        self._commit()
        sel = [self.rows[i] for i, v in enumerate(self._cvars) if v.get()]
        if not sel:
            messagebox.showwarning("Καμία επιλογή",
                "Τικάρισε τουλάχιστον ένα MRN.", parent=self.root)
            return
        self.result = sel; self.root.destroy()

    def _cancel(self):
        if messagebox.askyesno("Ακύρωση", "Σίγουρα θέλεις να ακυρώσεις;", parent=self.root):
            self.root.destroy()

    def run(self) -> list | None:
        self.root.mainloop(); return self.result

# ==============================================================================
# POPUP — Items popup (για 0832 με 2+ είδη)
# ==============================================================================

class ItemsPopup:
    """
    Εμφανίζεται ΜΟΝΟ για 0832 + ΔΑΣΜ_ΚΛ 76012080/76012030 με 2+ είδη.
    Δείχνει τα είδη από το XML και ζητά ΚΡΑΜΑ_1/ΒΑΡΟΣ_1 κλπ ανά είδος.
    Επιστρέφει dict {aa: {ΚΡΑΜΑ_1, ΒΑΡΟΣ_1, ΚΡΑΜΑ_2, ΒΑΡΟΣ_2, ΚΡΑΜΑ_3, ΒΑΡΟΣ_3}}
    ή None αν ακυρωθεί.
    """

    ITEM_SHOW_COLS = ["Α/Α", "ΔΑΣΜ_ΚΛ", "ΒΑΡΟΣ",
                      "ΚΡΑΜΑ_1", "ΒΑΡΟΣ_1", "ΚΡΑΜΑ_2", "ΒΑΡΟΣ_2", "ΚΡΑΜΑ_3", "ΒΑΡΟΣ_3"]
    ITEM_EDIT_COLS = {"ΚΡΑΜΑ_1", "ΒΑΡΟΣ_1", "ΚΡΑΜΑ_2", "ΒΑΡΟΣ_2", "ΚΡΑΜΑ_3", "ΒΑΡΟΣ_3"}

    def __init__(self, df_result: pd.DataFrame, mrn: str):
        self.df      = df_result.copy()
        self.mrn     = mrn
        self.result  = None
        self._edit_widget = None

        # Φτιάχνουμε dict με τιμές ανά Α/Α
        self.data = {}
        for _, row in self.df.iterrows():
            aa = str(row["Α/Α"])
            self.data[aa] = {
                "Α/Α":     aa,
                "ΔΑΣΜ_ΚΛ": str(row["ΔΑΣΜ_ΚΛ"]),
                "ΒΑΡΟΣ":   str(row["ΒΑΡΟΣ"]),
                "ΚΡΑΜΑ_1": "", "ΒΑΡΟΣ_1": "",
                "ΚΡΑΜΑ_2": "", "ΒΑΡΟΣ_2": "",
                "ΚΡΑΜΑ_3": "", "ΒΑΡΟΣ_3": "",
            }

        self.root = tk.Tk()
        self.root.title(f"Κράματα ανά Είδος — MRN: {mrn}")
        self.root.resizable(True, True)
        w, h = 1200, 600
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.root.configure(bg="#F5F4F0")
        self._build()
        self.root.lift()  # ← νέο
        self.root.focus_force()  # ← νέο
        self.root.attributes("-topmost", True)  # ← νέο

    def _build(self):
        r = self.root

        # Header
        hdr = tk.Frame(r, bg="#1A1A1A", height=52)
        hdr.pack(fill="x")
        tk.Label(hdr,
            text=f"  Ορισμός Κραμάτων ανά Είδος — MRN: {self.mrn}",
            bg="#1A1A1A", fg="#FFFFFF", font=("Consolas", 11), anchor="w"
        ).pack(side="left", padx=8, pady=14)

        # Hint
        hint = tk.Frame(r, bg="#F5F4F0")
        hint.pack(fill="x", padx=14, pady=(8, 2))
        tk.Label(hint,
            text="  Double-click για επεξεργασία   |   "
                 "Αν 1 κράμα: γράψε μόνο ΚΡΑΜΑ_1, άφησε ΒΑΡΟΣ_1 κενό   |   "
                 "Αν 2-3 κράματα: συμπλήρωσε ΚΡΑΜΑ + ΒΑΡΟΣ (άθροισμα = συνολικό βάρος είδους)",
            bg="#F5F4F0", fg="#6B6B65", font=("Consolas", 9), anchor="w"
        ).pack(side="left")

        # Treeview
        frm = tk.Frame(r, bg="#F5F4F0")
        frm.pack(fill="both", expand=True, padx=14, pady=(4, 6))

        style = ttk.Style(); style.theme_use("clam")
        style.configure("T2.Treeview",
            background="#FFFFFF", foreground="#1A1A1A", rowheight=28,
            fieldbackground="#FFFFFF", font=("Consolas", 10), borderwidth=0)
        style.configure("T2.Treeview.Heading",
            background="#E8E6DF", foreground="#3A3A36",
            font=("Consolas", 10, "bold"), relief="flat")
        style.map("T2.Treeview", background=[("selected", "#D4E8FF")])

        vsb = ttk.Scrollbar(frm, orient="vertical")
        hsb = ttk.Scrollbar(frm, orient="horizontal")
        self.tree = ttk.Treeview(frm, columns=self.ITEM_SHOW_COLS, show="headings",
            style="T2.Treeview", yscrollcommand=vsb.set, xscrollcommand=hsb.set,
            selectmode="browse")
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        col_w = {
            "Α/Α": 50, "ΔΑΣΜ_ΚΛ": 100, "ΒΑΡΟΣ": 90,
            "ΚΡΑΜΑ_1": 70, "ΒΑΡΟΣ_1": 80,
            "ΚΡΑΜΑ_2": 70, "ΒΑΡΟΣ_2": 80,
            "ΚΡΑΜΑ_3": 70, "ΒΑΡΟΣ_3": 80,
        }
        for c in self.ITEM_SHOW_COLS:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_w[c], anchor="w",
                stretch=(c in ("ΔΑΣΜ_ΚΛ", "ΒΑΡΟΣ")))

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.grid_rowconfigure(0, weight=1); frm.grid_columnconfigure(0, weight=1)

        self._iids = []
        for aa, d in self.data.items():
            vals = [d[c] for c in self.ITEM_SHOW_COLS]
            iid = self.tree.insert("", "end", iid=aa, values=vals)
            self._iids.append(iid)

        self.tree.bind("<Double-Button-1>", self._on_dbl)

        # Footer
        foot = tk.Frame(r, bg="#F5F4F0")
        foot.pack(fill="x", padx=14, pady=(0, 12))
        tk.Button(foot, text="Ακύρωση", font=("Consolas", 10),
            bg="#E8E6DF", fg="#6B6B65", relief="flat", padx=14, pady=7,
            cursor="hand2", command=self._cancel
        ).pack(side="right", padx=(6, 0))
        tk.Button(foot, text="OK — Συνέχεια", font=("Consolas", 10, "bold"),
            bg="#1A1A1A", fg="#FFFFFF", relief="flat", padx=14, pady=7,
            cursor="hand2", activebackground="#333", activeforeground="#FFF",
            command=self._ok
        ).pack(side="right")

    def _on_dbl(self, event):
        col_id = self.tree.identify_column(event.x)
        iid    = self.tree.identify_row(event.y)
        if not iid: return
        ci = int(col_id.replace("#", "")) - 1
        if ci < 0 or ci >= len(self.ITEM_SHOW_COLS): return
        cn = self.ITEM_SHOW_COLS[ci]
        if cn not in self.ITEM_EDIT_COLS: return
        self._commit()
        bbox = self.tree.bbox(iid, column=col_id)
        if not bbox: return
        x, y, w, h = bbox
        vals = list(self.tree.item(iid, "values"))
        e = tk.Entry(self.tree, font=("Consolas", 10), bg="#FFFDE7",
            fg="#1A1A1A", relief="solid", bd=1, insertbackground="#1A1A1A")
        e.place(x=x, y=y, width=w, height=h)
        e.insert(0, vals[ci]); e.select_range(0, "end"); e.focus_set()
        def commit(ev=None):
            vals[ci] = e.get().strip()
            self.tree.item(iid, values=vals)
            self.data[iid][cn] = vals[ci]
            e.destroy(); self._edit_widget = None
        e.bind("<Return>", commit); e.bind("<Tab>", commit)
        e.bind("<Escape>", lambda ev: e.destroy()); e.bind("<FocusOut>", commit)
        self._edit_widget = e

    def _commit(self):
        if self._edit_widget:
            try: self._edit_widget.event_generate("<FocusOut>")
            except: pass

    def _ok(self):
        self._commit()
        # Validation ανά είδος
        errors = []
        for aa, d in self.data.items():
            kr1 = d["ΚΡΑΜΑ_1"].strip()
            if not kr1:
                errors.append(f"Α/Α {aa}: ΚΡΑΜΑ_1 είναι κενό!")
                continue

            # Αν έχει ΚΡΑΜΑ_2, πρέπει να έχει και ΒΑΡΟΣ_1 + ΒΑΡΟΣ_2
            kr2 = d["ΚΡΑΜΑ_2"].strip()
            kr3 = d["ΚΡΑΜΑ_3"].strip()
            has_multiple = bool(kr2)

            if has_multiple:
                # Validation αθροίσματος
                try:
                    # Παίρνουμε το συνολικό βάρος (αντικατάσταση κόμμα→τελεία)
                    total = float(d["ΒΑΡΟΣ"].replace(",", "."))
                    b1 = float(d["ΒΑΡΟΣ_1"].replace(",", ".")) if d["ΒΑΡΟΣ_1"].strip() else 0
                    b2 = float(d["ΒΑΡΟΣ_2"].replace(",", ".")) if d["ΒΑΡΟΣ_2"].strip() else 0
                    b3 = float(d["ΒΑΡΟΣ_3"].replace(",", ".")) if d["ΒΑΡΟΣ_3"].strip() else 0
                    s  = b1 + b2 + (b3 if kr3 else 0)
                    if abs(s - total) > 0.01:
                        errors.append(
                            f"Α/Α {aa}: άθροισμα βαρών ({s}) ≠ συνολικό βάρος ({total})!")
                except ValueError:
                    errors.append(f"Α/Α {aa}: μη έγκυρες τιμές βάρους!")

        if errors:
            messagebox.showerror("Σφάλματα",
                "\n".join(errors), parent=self.root)
            return

        self.result = self.data
        self.root.destroy()

    def _cancel(self):
        if messagebox.askyesno("Ακύρωση", "Σίγουρα θέλεις να ακυρώσεις;", parent=self.root):
            self.root.destroy()

    def run(self) -> dict | None:
        self.root.mainloop(); return self.result

# ==============================================================================
# HELPER — Expand 0832 rows με κράματα
# ==============================================================================

KRAMMA_DASMOS_KL = {"76012080", "76012030", "76011010"}

def is_0832_kramma(row) -> bool:
    """Ελέγχει αν η γραμμή χρειάζεται χειροκίνητο κράμα (0832 + συγκεκριμένη ΔΑΣΜ_ΚΛ)."""
    return (str(row.get("ΤΕΛΩΝ", "")) == "0832" and
            str(row.get("ΔΑΣΜ_ΚΛ", "")) in KRAMMA_DASMOS_KL)


def expand_krammata(df_result: pd.DataFrame, krammata: dict) -> pd.DataFrame:
    """
    Σπάει κάθε είδος σε 1-3 γραμμές ανάλογα με τα κράματα.

    krammata: dict με keys = str(Α/Α), values = dict με ΚΡΑΜΑ_1/ΒΑΡΟΣ_1 κλπ
              ή απευθείας από το 1ο popup (ένα dict για όλα τα είδη)

    Κανόνες:
    - Αν ΚΡΑΜΑ_1 μόνο (ΒΑΡΟΣ_1 κενό) → 1 γραμμή με το συνολικό βάρος
    - Αν ΚΡΑΜΑ_1 + ΒΑΡΟΣ_1 + ΚΡΑΜΑ_2 + ΒΑΡΟΣ_2 → 2 γραμμές με τα αντίστοιχα βάρη
    - Αν και ΚΡΑΜΑ_3 → 3 γραμμές
    """
    new_rows = []
    for _, row in df_result.iterrows():
        aa  = str(row["Α/Α"])
        # Αν δεν είναι 0832 kramma → αφήνουμε ως έχει
        if not is_0832_kramma(row):
            new_rows.append(row.to_dict())
            continue

        d = krammata.get(aa, {})
        kr1 = d.get("ΚΡΑΜΑ_1", "").strip()
        vr1 = d.get("ΒΑΡΟΣ_1", "").strip()
        kr2 = d.get("ΚΡΑΜΑ_2", "").strip()
        vr2 = d.get("ΒΑΡΟΣ_2", "").strip()
        kr3 = d.get("ΚΡΑΜΑ_3", "").strip()
        vr3 = d.get("ΒΑΡΟΣ_3", "").strip()

        if not kr1:
            # Δεν έχει κράμα → αφήνουμε ως έχει (θα ρωτήσει popup_input μέσα στο sap_entry)
            new_rows.append(row.to_dict())
            continue

        # Γραμμή 1
        r1 = row.to_dict()
        r1["ΚΡΑΜΑ"] = kr1
        # Αν ΒΑΡΟΣ_1 κενό → παίρνει το συνολικό βάρος (μόνο 1 κράμα)
        r1["ΒΑΡΟΣ"] = vr1 if vr1 else row["ΒΑΡΟΣ"]
        new_rows.append(r1)

        # Γραμμή 2 (αν υπάρχει ΚΡΑΜΑ_2)
        if kr2:
            r2 = row.to_dict()
            r2["ΚΡΑΜΑ"] = kr2
            r2["ΒΑΡΟΣ"] = vr2
            new_rows.append(r2)

        # Γραμμή 3 (αν υπάρχει ΚΡΑΜΑ_3)
        if kr3:
            r3 = row.to_dict()
            r3["ΚΡΑΜΑ"] = kr3
            r3["ΒΑΡΟΣ"] = vr3
            new_rows.append(r3)

    return pd.DataFrame(new_rows).reset_index(drop=True)

# ==============================================================================
# ΦΑΣΗ Β.1 — Download XML
# ==============================================================================

def phase_b_download_xml(mrn: str, driver):
    SAVE_FOLDER.mkdir(parents=True, exist_ok=True)
    for f in SAVE_FOLDER.glob("*.xml"):
        try: f.unlink()
        except: pass

    # Έλεγξε αν ο browser είναι ζωντανός
    try:
        _ = driver.current_url
    except:
        print("    Browser έκλεισε — επανασύνδεση ICISnet...")
        driver = make_chrome(download_folder=SAVE_FOLDER)
        wait   = WebDriverWait(driver, 30)
        driver.get("https://www1.gsis.gr/icisnet/itrader/common/home.jsf")
        wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(ICISNET_USER)
        driver.find_element(By.NAME, "password").send_keys(ICISNET_PASS)
        driver.find_element(By.NAME, "btn_login").click()
        time.sleep(3)

    # Επιστροφή στο menu — ← ΕΔΩ
    driver.get("https://www1.gsis.gr/icisnet/itrader/common/home.jsf")
    driver.maximize_window()
    time.sleep(3)

    wait = WebDriverWait(driver, 30)
    print(f"    ICISnet download XML για MRN={mrn}...")

    clicked = False
    for _ in range(5):
        try:
            driver.execute_script("arguments[0].click();",
                driver.find_element(By.ID, "tablehidecontentForm:actions_message_search"))
            clicked = True; break
        except: time.sleep(1)
    if not clicked:
        for _ in range(5):
            try:
                driver.execute_script("arguments[0].click();",
                    driver.find_element(By.ID, "contentForm:actions_message_search"))
                time.sleep(2)
                driver.execute_script("arguments[0].click();",
                    driver.find_element(By.ID, "tablehidecontentForm:actions_message_search"))
                clicked = True; break
            except: time.sleep(1)
    if not clicked: raise Exception("Menu navigation failed")

    wait.until(EC.presence_of_element_located((By.ID, "contentForm:mrn-arc-nid")))
    inp = wait.until(EC.element_to_be_clickable((By.ID, "contentForm:mrn-arc-nid")))
    inp.clear(); inp.send_keys(mrn)
    safe_select(wait, "contentForm:domain",       "Εισαγωγές"); time.sleep(2)
    safe_select(wait, "contentForm:message_type", "ID29")
    safe_select(wait, "contentForm:search_scope", "Συναλλασσόμενος")

    driver.execute_script("arguments[0].click();",
        wait.until(EC.element_to_be_clickable((By.ID, "contentForm:submitFormButton"))))
    wait.until(EC.presence_of_element_located((By.ID, "contentForm:search_results")))

    driver.execute_script("arguments[0].click();",
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "ID29"))))

    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[contains(text(),'Ενέργειες')]"))).click()
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[contains(text(),'Αποθήκευση ως XML')]"))).click()

    for _ in range(20):
        time.sleep(1)
        xmls = list(SAVE_FOLDER.glob("*.xml"))
        if xmls:
            latest = max(xmls, key=lambda f: f.stat().st_ctime)
            if XML_PATH.exists(): XML_PATH.unlink()
            latest.rename(XML_PATH)
            print(f"    XML saved: {XML_PATH.name}")
            driver.minimize_window()
            return driver

    raise Exception("XML download timed out")

# ==============================================================================
# ΦΑΣΗ Β.2 — Parse XML
# ==============================================================================

def phase_b_parse_xml() -> pd.DataFrame:
    import xml.etree.ElementTree as ET
    tree = ET.parse(XML_PATH); root = tree.getroot()

    def fmt(d): return d.strftime("%d.%m.%Y")
    def to_comma(val):
        if val is None or val == "": return ""
        try:
            f = float(val)
            s = str(int(f)) if f == int(f) else f"{f:.2f}"
            return s.replace(".", ",")
        except: return str(val).replace(".", ",")

    def to_comma_rate(val):
        """Για ισοτιμία — max 5 δεκαδικά."""
        if val is None or val == "": return ""
        try:
            f = float(val)
            s = str(int(f)) if f == int(f) else f"{f:.5f}".rstrip("0")
            return s.replace(".", ",")
        except:
            return str(val).replace(".", ",")

    hea = root.find("HEAHEA")
    if hea is None: raise ValueError("HEAHEA tag not found in XML")

    mrn         = hea.findtext("DocNumHEA5", "")
    typ_dec     = hea.findtext("TypOfDecHEA24", "")
    typ_dec_b12 = hea.findtext("TypOfDecBx12HEA651", "")
    cou_dis     = hea.findtext("CouOfDisCodHEA55", "")
    dec_date    = datetime.strptime(hea.findtext("DecDatHEA383", ""), "%Y%m%d").date()
    typos       = typ_dec + typ_dec_b12

    hl  = dec_date + relativedelta(months=32) - relativedelta(days=1)
    hl2 = hl + relativedelta(months=-26)
    hp  = hl2 + relativedelta(months=1)

    telwneio = root.find("IMPCUSOFF").findtext("RefNumIMPCUSOFF", "")[-4:]
    prom     = root.find("TRACONCO1").findtext("NamCO17", "").strip()
    oroi     = root.find("DELTER").findtext("IncCodTDL1", "")
    tradat   = root.find("TRADAT")
    nom_isot = tradat.findtext("CurTRD1", "")
    isot     = tradat.findtext("ExcRatTRD1", "1")

    rows = []
    for item in root.findall("GOOITEGDS"):
        aa       = item.findtext("IteNumGDS7", "")
        desc     = item.findtext("GooDesGDS23", "")
        net_mas  = item.findtext("NetMasGDS48", "")
        pro_req  = item.findtext("ProReqGDI1", "")
        pre_pro  = item.findtext("PreProGDI1", "")
        nom_stat = item.findtext("StaValCurGDI1", "")
        cou_ori  = item.findtext("CouOfOriGDI1", "")
        pro_pri  = item.findtext("ProPri4002", "")
        sta_val  = item.findtext("StaValAmoGDI1", "")

        kramma = ""
        if "SLABS " in desc:
            kramma = desc.split("SLABS ", 1)[1].split(" ", 1)[0][:4]

        tel_kath = pro_req + pre_pro
        rel      = item.find("REL800")
        rel_cod  = rel.findtext("RelRelCod02", "") if rel is not None else ""

        if rel_cod == "X16":              kath = 3
        elif tel_kath.startswith("51"):   kath = 2
        else:                             kath = 5

        xwra   = cou_dis.replace("XS", "RS") if cou_ori == "EU" else cou_ori
        ent_ag = ""
        for doc in item.findall("PRODOCDC2"):
            di = doc.findtext("DocInfDC1008", "")
            if di: ent_ag = di.replace("/", "000").replace("-", "000"); break

        comcod    = item.find("COMCODGODITM")
        dasmos_kl = comcod.findtext("ComNomCMD1", "") if comcod is not None else ""

        taxes = [(float(t.findtext("RatOfTaxCTX1", "0")),
                  float(t.findtext("AmoOfTaxTCL1", "0")))
                 for t in item.findall("CALTAXGOD")]
        if taxes:
            rates = [t[0] for t in taxes]; amos = [t[1] for t in taxes]
            sd = min(rates); da = min(amos); sf = max(rates); fa = max(amos)
            df_val = None if da == fa else da; sf_fin = sf if sf else 24.0
        else:
            sd = da = fa = None; df_val = None; sf_fin = 24.0

        rows.append({
            "ΤΕΛΩΝ": telwneio, "ΤΥΠΟΣ": typos, "MRN": mrn, "ΗΜΕΡ": fmt(dec_date),
            "ΔΑΣΜ_ΚΛ": dasmos_kl, "Α/Α": int(aa) if aa.isdigit() else aa,
            "ΧΩΡΑ": xwra, "ΚΡΑΜΑ": kramma, "ΚΑΘ": kath, "ΤΕΛ_ΚΑΘ": tel_kath,
            "ΗΜΕΡ_ΛΗΞΗΣ": fmt(hl), "ΗΜΕΡ_ΛΗΞΗΣ2": fmt(hl2),
            "ΒΑΡΟΣ": to_comma(net_mas), "X16": rel_cod, "ΝΟΜ_ΣΤΑΤ": nom_stat,
            "ΣΤΑΤ_ΑΞΙΑ": f"{float(sta_val):.2f}".replace(".", ",") if sta_val else "",
            "ΠΡΟΜ": prom, "ΟΡΟΙ": oroi,
            "ΤΙΜΗ": f"{float(pro_pri):.2f}".replace(".", ",") if pro_pri else "",
            "ΝΟΜ_ΙΣΟΤ": nom_isot, "ΙΣΟΤ": to_comma_rate(isot),
            "ΣΥΝΤ_ΔΑΣΜ": to_comma(sd), "ΔΑΣΜ": to_comma(df_val),
            "ΣΥΝΤ_ΦΠΑ": to_comma(sf_fin), "ΦΠΑ": to_comma(fa),
            "ΕΝΤ_ΑΓ": ent_ag, "ΗΜΕΡ_ΠΡΟΘ": fmt(hp),
        })

    cols = ["ΤΕΛΩΝ","ΤΥΠΟΣ","MRN","ΗΜΕΡ","ΔΑΣΜ_ΚΛ","Α/Α","ΧΩΡΑ","ΚΡΑΜΑ","ΚΑΘ","ΤΕΛ_ΚΑΘ",
            "ΗΜΕΡ_ΛΗΞΗΣ","ΗΜΕΡ_ΛΗΞΗΣ2","ΒΑΡΟΣ","X16","ΝΟΜ_ΣΤΑΤ","ΣΤΑΤ_ΑΞΙΑ","ΠΡΟΜ","ΟΡΟΙ",
            "ΤΙΜΗ","ΝΟΜ_ΙΣΟΤ","ΙΣΟΤ","ΣΥΝΤ_ΔΑΣΜ","ΔΑΣΜ","ΣΥΝΤ_ΦΠΑ","ΦΠΑ","ΕΝΤ_ΑΓ","ΗΜΕΡ_ΠΡΟΘ"]
    return pd.DataFrame(rows)[cols]

# ==============================================================================
# ΦΑΣΗ Β.3 — SAP entry
# ==============================================================================

def get_kramma(dasmos_kl: str, synt_dasm: str):
    if dasmos_kl.startswith("760200"):
        return "SCRAP-ΦΥΡΑ" if synt_dasm == "24" else "SCRAP"
    m = {
        "76011090": "1XXX", "76012080": "AL", "76012030": "AL", "76011010": "AL",
        "81041100": "MG",   "81110011": "MN",
        "72052900": "FE",   "72069000": "FE", "72029980": "FE",
    }
    return m.get(dasmos_kl)


def focus_sap():
    try:
        Application(backend="uia")\
            .connect(title_re=".*Νέες.*Καταχωρίσεων.*", timeout=5)\
            .window(title_re=".*Νέες.*Καταχωρίσεων.*").set_focus()
        time.sleep(0.3)
    except:
        try:
            Application(backend="uia")\
                .connect(class_name="SAP_FRONTEND_SESSION", timeout=5)\
                .window(class_name="SAP_FRONTEND_SESSION").set_focus()
            time.sleep(0.3)
        except: pass

SAP_WIN_TITLE = "Νέες Καταχωρίσεις: Λεπτομέρειες Προστιθέμενων Καταχωρίσεων"


def get_sap_win(timeout=10):
    app = Application(backend="uia").connect(
        class_name="SAP_FRONTEND_SESSION", title=SAP_WIN_TITLE, timeout=timeout)
    return app.window(class_name="SAP_FRONTEND_SESSION", title=SAP_WIN_TITLE)


def sap_entry(row, prot: str, sap_running: bool = False):
    tl  = str(row["ΤΕΛΩΝ"]);  ty  = str(row["ΤΥΠΟΣ"]); mn  = str(row["MRN"])
    im  = str(row["ΗΜΕΡ"]);   dk  = str(row["ΔΑΣΜ_ΚΛ"]); aa = str(row["Α/Α"])
    xw  = str(row["ΧΩΡΑ"]);   x16 = str(row["X16"]); tk_ = str(row["ΤΕΛ_ΚΑΘ"])
    il  = str(row["ΗΜΕΡ_ΛΗΞΗΣ"]); il2 = str(row["ΗΜΕΡ_ΛΗΞΗΣ2"])
    vr  = str(row["ΒΑΡΟΣ"]);  sa  = str(row["ΣΤΑΤ_ΑΞΙΑ"]); ns = str(row["ΝΟΜ_ΣΤΑΤ"])
    pr  = str(row["ΠΡΟΜ"]);   or_ = str(row["ΟΡΟΙ"]); ti = str(row["ΤΙΜΗ"])
    ni  = str(row["ΝΟΜ_ΙΣΟΤ"]); is_ = str(row["ΙΣΟΤ"]); sd = str(row["ΣΥΝΤ_ΔΑΣΜ"])
    da  = str(row["ΔΑΣΜ"]);   sf  = str(row["ΣΥΝΤ_ΦΠΑ"]); fp = str(row["ΦΠΑ"])
    ea  = str(row["ΕΝΤ_ΑΓ"]).strip(); ip = str(row["ΗΜΕΡ_ΠΡΟΘ"])

    if x16 == "X16":             kath = 3
    elif tk_ == "5111":          kath = 12
    elif tk_.startswith("5"):    kath = 2
    else:                        kath = 5

    # ── ΚΡΑΜΑ: για 0832+ΔΑΣΜ_ΚΛ παίρνει από τη γραμμή, αλλιώς get_kramma ──
    popup_kr = str(row.get("ΚΡΑΜΑ", "")).strip()
    if popup_kr and popup_kr not in ("nan", ""):
        kr = popup_kr  # popup έχει προτεραιότητα πάντα
    elif is_0832_kramma(row):
        kr = popup_input("ΚΡΑΜΑ", f"ΤΕΛΩΝ=0832 | {dk}\nΕισάγετε ΚΡΑΜΑ:")
    else:
        kr = get_kramma(dk, sd)
        if kr is None:
            kr = popup_input("ΚΡΑΜΑ", f"Εισάγετε ΚΡΑΜΑ για {dk}:")

    if not prot or prot in ("", "nan"):
        prot = popup_input("ΠΡΟΤΙΜΗΣΗ", f"MRN:{mn}|Α/Α:{aa}\nΕισάγετε ΠΡΟΤΙΜΗΣΗ:")
        if not prot: raise ValueError("ΠΡΟΤΙΜΗΣΗ κενή")

    if sd != "24" and ea in ("", "nan", "None"):

        raw = popup_input("ΕΝΤΟΛΗ ΑΓΟΡΑΣ", f"MRN:{mn}\n5 ψηφία ΕΝΤ_ΑΓ (ή NO):")
        if raw:
            ea = "NO" if raw.upper() == "NO" else f"41000{raw.strip()}"

    print(f"    Α/Α={aa} | ΚΑΘ={kath} | ΚΡΑΜΑ={kr} | ΒΑΡΟΣ={vr} | ΤΕΛΩΝ={tl}")

    if not sap_running:
        close_sap(); time.sleep(2)
        setup_keyboard()
        subprocess.Popen(f'start "" "{SAP_FILE_B}"', shell=True); time.sleep(2.5)
        ap = Application(backend="uia").connect(title="ZELVMM_IMP_1", timeout=60)
        wn = ap.window(title="ZELVMM_IMP_1")
        wn.child_window(auto_id="1004", control_type="Edit")\
          .wait("exists enabled visible", timeout=30).set_text(SAP_USER)
        wn.child_window(auto_id="1005", control_type="Edit").set_text(SAP_PASS)
        wn.child_window(title="Εισ.σε Σύστ.", control_type="Button").click_input()
        time.sleep(3)


        # Στέλνουμε τα TABs απευθείας στο "Επιλογή Πεδίου"
        # Επιλογή Πεδίου
        print("Ψάχνω Επιλογή Πεδίου...")
        while True:
            try:
                dlg = Application(backend="uia").connect(title="Επιλογή Πεδίου", timeout=2)
                win_dlg = dlg.window(title="Επιλογή Πεδίου")
                win_dlg.set_focus()
                hwnd = win_dlg.handle
                fore_hwnd = ctypes.windll.user32.GetForegroundWindow()
                fore_tid = ctypes.windll.user32.GetWindowThreadProcessId(fore_hwnd, None)
                this_tid = ctypes.windll.kernel32.GetCurrentThreadId()
                ctypes.windll.user32.AttachThreadInput(this_tid, fore_tid, True)
                ctypes.windll.user32.SetForegroundWindow(hwnd)
                ctypes.windll.user32.AttachThreadInput(this_tid, fore_tid, False)
                time.sleep(1)
                break
            except:
                time.sleep(0.5)

        pyautogui.hotkey("tab")
        pyautogui.hotkey("tab")
        pyautogui.hotkey("tab")
        pyautogui.hotkey("tab")
        pyautogui.hotkey("space")
        pyautogui.hotkey("enter")

        time.sleep(1)

        send_keys("1100{ENTER}", pause=0.05)
        time.sleep(1)
        print("Έστειλε την εταιρεία")
    for _ in range(10):
        try:
            ap2 = Application(backend="uia").connect(title_re=".*Επισκόπηση.*", timeout=5)
            wn2 = ap2.window(title_re=".*Επισκόπηση.*"); wn2.maximize()
            try:
                Application(backend="uia").connect(title="Πληροφορίες", timeout=2)
                wn2.close(); time.sleep(1)
                setup_keyboard()
                subprocess.Popen(f'start "" "{SAP_FILE_B}"', shell=True); time.sleep(2.5)
                wr = Application(backend="uia").connect(title="ZELVMM_IMP_1", timeout=60)\
                       .window(title="ZELVMM_IMP_1")
                wr.child_window(auto_id="1004", control_type="Edit")\
                  .wait("exists enabled visible", timeout=30).set_text(SAP_USER)
                wr.child_window(auto_id="1005", control_type="Edit").set_text(SAP_PASS)
                wr.child_window(title="Εισ.σε Σύστ.", control_type="Button").click_input()
                time.sleep(3)
                # Επιλογή Πεδίου
                print("Ψάχνω Επιλογή Πεδίου...")
                while True:
                    try:
                        dlg = Application(backend="uia").connect(title="Επιλογή Πεδίου", timeout=2)
                        win_dlg = dlg.window(title="Επιλογή Πεδίου")
                        win_dlg.set_focus()
                        hwnd = win_dlg.handle
                        fore_hwnd = ctypes.windll.user32.GetForegroundWindow()
                        fore_tid = ctypes.windll.user32.GetWindowThreadProcessId(fore_hwnd, None)
                        this_tid = ctypes.windll.kernel32.GetCurrentThreadId()
                        ctypes.windll.user32.AttachThreadInput(this_tid, fore_tid, True)
                        ctypes.windll.user32.SetForegroundWindow(hwnd)
                        ctypes.windll.user32.AttachThreadInput(this_tid, fore_tid, False)
                        time.sleep(1)
                        break
                    except:
                        time.sleep(0.5)
                pyautogui.hotkey("tab")
                pyautogui.hotkey("tab")
                pyautogui.hotkey("tab")
                pyautogui.hotkey("tab")
                pyautogui.hotkey("space")
                pyautogui.hotkey("enter")

                time.sleep(1)
                # 1100 + ENTER
                print("Sending company code 1100...")
                send_keys("1100{ENTER}", pause=0.05)
                time.sleep(1)
                print("Done!")
            except: pass
            break
        except: time.sleep(1)

    # Step 1: Wait for overview window
    while True:
        try:
            ap2 = Application(backend="uia").connect(title_re=".*Επισκόπηση.*", timeout=5)
            wn2 = ap2.window(title_re=".*Επισκόπηση.*")
            wn2.maximize()
            break
        except:
            time.sleep(2)

    # Step 2: Click Change — if info dialog opens, click Continue and retry
    while True:
        try:
            wn2.child_window(title="Εμφάνιση -> Αλλαγή", control_type="Button") \
                .wait("exists enabled visible", timeout=30).click_input()
            time.sleep(2)

            # Check if info dialog opened
            try:
                wn2.child_window(class_name_re=".*:00000000:00000000", title="AppToolbar") \
                    .child_window(title="Συνέχεια", class_name="Button") \
                    .wait("exists enabled visible", timeout=3).click_input()
                print("  Another user active — Continue and retry...")
                time.sleep(1)
                continue
            except:
                pass

            # Step 3: Click New Entries — wait until available
            while True:
                try:
                    ap3 = Application(backend="uia").connect(title_re=".*Αλλαγή.*Επισκόπηση.*", timeout=5)
                    wn3 = ap3.window(title_re=".*Αλλαγή.*Επισκόπηση.*")
                    wn3.child_window(title="Νέες Καταχωρίσεις", control_type="Button") \
                        .wait("exists enabled visible", timeout=30).click_input()
                    time.sleep(1)
                    print("  New Entries OK!")
                    # Maximise the new entries window
                    while True:
                        try:
                            ap4 = Application(backend="uia").connect(title=SAP_WIN_TITLE, timeout=10)
                            ap4.window(title=SAP_WIN_TITLE).maximize()
                            break
                        except:
                            time.sleep(1)
                    break
                except:
                    time.sleep(1)
            break
        except:
            time.sleep(1)


    if tl != "0832":
        send_keys(f"{tl}{{TAB}}{tl}{{TAB}}{ty}{{TAB}}{{TAB}}"
                  f"{mn}{{TAB}}{mn}{{TAB}}{{TAB}}"
                  f"{im}{{TAB}}{im}{{TAB}}{{TAB}}"
                  f"{dk}{{TAB}}{{TAB}}1100{{TAB}}{aa}{{TAB}}{xw}{{TAB}}", pause=0.05)
    else:
        send_keys(f"{tl}{{TAB}}{{TAB}}{ty}{{TAB}}{{TAB}}"
                  f"{mn}{{TAB}}{{TAB}}{{TAB}}"
                  f"{im}{{TAB}}{{TAB}}{{TAB}}"
                  f"{dk}{{TAB}}{{TAB}}1100{{TAB}}{aa}{{TAB}}{xw}{{TAB}}", pause=0.05)

    send_keys(f"{kath}{{TAB}}{tk_}{{TAB}}", pause=0.05)

    is_swan = (kath == 3 and tk_ == "6121" and pr == "SUPPLIER_NAME_PASSIVE" and dk == "76012080")
    if is_swan:
        send_keys(f"{il}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}"
                  f"AUTH_NUMBER_PASSIVE{{TAB}}"
                  f"{vr}{{TAB}}KG{{TAB}}{sa}{{TAB}}{ns}{{TAB}}"
                  f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{SPACE}}{{TAB}}{{TAB}}", pause=0.05)
        send_keys("SUPPLIER_NAME_PASSIVE - ΠΑΘΗΤΙΚΗ ΤΕΛ.", with_spaces=True, pause=0.05)
        send_keys("{TAB}{TAB}{TAB}{TAB}{TAB}{TAB}{TAB}{TAB}AUTH_DECISION_X16{TAB}{TAB}AL{TAB}{TAB}{TAB}{TAB}{TAB}{TAB}\n",pause=0.05)

        if sd == "24":
            send_keys(f"{prot}{{TAB}}{{TAB}}{or_}{{TAB}}{ti}{{TAB}}{{TAB}}"
                      f"{ni}{{TAB}}{{TAB}}{is_}{{TAB}}{{TAB}}"
                      f"{sa}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}"
                      f"{sf}{{TAB}}{{TAB}}{fp}", pause=0.05)
        else:
            send_keys(f"{prot}{{TAB}}{{TAB}}{or_}{{TAB}}{ti}{{TAB}}{{TAB}}"
                      f"{ni}{{TAB}}{{TAB}}{is_}{{TAB}}{{TAB}}"
                      f"{sa}{{TAB}}{{TAB}}{sd}{{TAB}}{da}{{TAB}}{{TAB}}"
                      f"{sf}{{TAB}}{{TAB}}{fp}", pause=0.05)
            if ea.upper() == "NO":
                send_keys("{TAB}{TAB}{ENTER}", pause=0.05)
            else:
                send_keys(f"{{TAB}}{{TAB}}{ea}{{ENTER}}", pause=0.05)

        send_keys("{ENTER}")
        time.sleep(0.5)
        win = get_sap_win()
        win.set_focus()
        time.sleep(0.5)
        send_keys("{F11}")

        time.sleep(0.5); return

    if kath == 3:         send_keys(f"{il}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}", pause=0.05)
    elif kath in (5, 12): send_keys("{TAB}{TAB}{TAB}{TAB}{TAB}", pause=0.05)
    else:                 send_keys(f"{il2}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}AUTH_NUMBER_IP", pause=0.05)

    send_keys(f"{{TAB}}{vr}{{TAB}}KG{{TAB}}{sa}{{TAB}}{ns}{{TAB}}{{TAB}}", pause=0.05)

    if sd == "24":
        focus_sap(); send_keys("+%(1)", pause=0.05); time.sleep(0.5)
        pr = f"ΦΥΡΑ - ΕΚΚΑΘΑΡΙΣΗ {im}"

    is_c = ty.upper().endswith("C")
    if is_c:
        if kath == 2:
            send_keys(f"AUTH_NUMBER_IP_CGU{{TAB}}{{TAB}}AUTH_NUMBER_BOND_IP{{TAB}}{{SPACE}}{{TAB}}"
                      f"AUTH_NUMBER_SASP_IP{{TAB}}", pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{ip}{{TAB}}{{TAB}}AUTH_DECISION_IP", pause=0.05)
        elif kath == 3:
            send_keys(f"AUTH_NUMBER_X16_CGU{{TAB}}{{TAB}}AUTH_NUMBER_BOND_X16{{TAB}}{{SPACE}}{{TAB}}"
                      f"AUTH_NUMBER_SASP_X16{{TAB}}", pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}AUTH_DECISION_X16", pause=0.05)
        elif kath == 5:
            send_keys(f"AUTH_NUMBER_X16_CGU{{TAB}}{{TAB}}AUTH_NUMBER_BOND_X16{{TAB}}{{SPACE}}{{TAB}}"
                      f"AUTH_NUMBER_SASP_X16{{TAB}}", pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}.", pause=0.05)
    else:
        if kath == 2:
            send_keys(f"AUTH_NUMBER_IP_CGU{{TAB}}{{TAB}}AUTH_NUMBER_BOND_IP{{TAB}}{{SPACE}}{{TAB}}{{TAB}}",
                      pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{ip}{{TAB}}{{TAB}}AUTH_DECISION_IP", pause=0.05)
        elif kath == 3:
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{SPACE}}{{TAB}}{{TAB}}", pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}AUTH_DECISION_X16", pause=0.05)
        elif kath in (5, 12):
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{SPACE}}{{TAB}}{{TAB}}", pause=0.05)
            send_keys(pr, with_spaces=True, pause=0.05)
            send_keys(f"{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}.", pause=0.05)

    send_keys(f"{{TAB}}{{TAB}}{kr}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}", pause=0.05)

    if sd == "24":
        send_keys(f"{prot}{{TAB}}{{TAB}}{or_}{{TAB}}{ti}{{TAB}}{{TAB}}"
                  f"{ni}{{TAB}}{{TAB}}{is_}{{TAB}}{{TAB}}"
                  f"{sa}{{TAB}}{{TAB}}{{TAB}}{{TAB}}{{TAB}}"
                  f"{sf}{{TAB}}{{TAB}}{fp}", pause=0.05)
    else:
        send_keys(f"{prot}{{TAB}}{{TAB}}{or_}{{TAB}}{ti}{{TAB}}{{TAB}}"
                  f"{ni}{{TAB}}{{TAB}}{is_}{{TAB}}{{TAB}}"
                  f"{sa}{{TAB}}{{TAB}}{sd}{{TAB}}{da}{{TAB}}{{TAB}}"
                  f"{sf}{{TAB}}{{TAB}}{fp}", pause=0.05)
        if ea.upper() == "NO": send_keys("{TAB}{TAB}{ENTER}", pause=0.05)
        else: send_keys(f"{{TAB}}{{TAB}}{ea}{{ENTER}}", pause=0.05)

    send_keys("{ENTER}")
    time.sleep(0.5)
    win = get_sap_win()
    win.set_focus()
    time.sleep(0.5)
    send_keys("{F11}")
    time.sleep(0.5)

# ==============================================================================
# ΦΑΣΗ Β.4 — SAP attach PDF
# ==============================================================================

def sap_attach(mrn: str, typos: str):
    pdf_name = f"{mrn} {typos}.pdf"
    full_path = str(SAP_GUI_DIR / pdf_name)
    print(f"    Attaching: {pdf_name}")

    for attempt in range(5):
        try:
            wn = get_sap_win(); wn.set_focus(); time.sleep(0.3)
            split = (wn.child_window(class_name="GOSContainer Class")
                       .child_window(class_name="Shell Window Class", found_index=0)
                       .child_window(class_name_re="ATL:.*", found_index=0)
                       .child_window(class_name="SysPager")
                       .child_window(class_name="ToolbarWindow32")
                       .child_window(control_type="SplitButton"))
            split.click_input(); time.sleep(0.7); break
        except Exception as e:
            if attempt == 4: raise RuntimeError(f"GOS SplitButton: {e}")
            time.sleep(3)

    for attempt in range(5):
        try:
            wn = get_sap_win()
            zg = wn.child_window(class_name="DialogBox Container Class", title_re="ZGOS_ZIMP1.*")
            tb = (zg.child_window(class_name="Shell Window Class", found_index=0)
                    .child_window(class_name_re="ATL:.*", found_index=0)
                    .child_window(class_name="SysPager")
                    .child_window(class_name="ToolbarWindow32"))
            tb.children()[0].click_input(); time.sleep(0.7); break
        except Exception as e:
            if attempt == 4: raise RuntimeError(f"Button[0]: {e}")
            time.sleep(3)

    for attempt in range(5):
        try:
            ct = Application(backend="uia").connect(title="Context", class_name="#32768", timeout=5)
            ct.window(title="Context", class_name="#32768")\
              .child_window(title="Δημιουργία Προσάρτησης", control_type="MenuItem").click_input()
            time.sleep(1); break
        except Exception as e:
            if attempt == 4: raise RuntimeError(f"MenuItem: {e}")
            time.sleep(2)

    def get_file_dlg():
        wn = get_sap_win()
        zg = wn.child_window(class_name="DialogBox Container Class", title_re="ZGOS_ZIMP1.*")
        return zg.child_window(class_name="#32770", title="Εισαγωγή αρχείου")

    for attempt in range(5):
        try:
            dlg = get_file_dlg()
            dlg.child_window(class_name="Edit", title="File name:").set_edit_text(full_path)
            time.sleep(0.3)
            dlg.child_window(class_name="Button", title="Open").click_input(); break
        except Exception as e:
            if attempt == 4: raise RuntimeError(f"File/Open: {e}")
            time.sleep(4)

    time.sleep(4); print(f"    Attached OK: {pdf_name}")

# ==============================================================================
# PROCESS ONE MRN
# ==============================================================================

def process_mrn(row: dict, driver) -> bool:
    mrn  = str(row.get("MRN",  "")).strip()
    prot = str(row.get("PROT", "")).strip()
    pdf  = str(row.get("PDF",  "")).strip()
    pdf1 = str(row.get("PDF.1","")).strip()

    prot = "" if prot in ("nan","None") else prot
    pdf  = "" if pdf  in ("nan","None") else pdf
    pdf1 = "" if pdf1 in ("nan","None") else pdf1
    if prot and prot.replace(".", "").isdigit():
        prot = str(int(float(prot)))

    # Alloy data from the approval popup (for customs office 0832)
    krammata_from_popup = {
        "ΚΡΑΜΑ_1": str(row.get("ΚΡΑΜΑ_1", "")).strip(),
        "ΒΑΡΟΣ_1":  str(row.get("ΒΑΡΟΣ_1",  "")).strip(),
        "ΚΡΑΜΑ_2": str(row.get("ΚΡΑΜΑ_2", "")).strip(),
        "ΒΑΡΟΣ_2":  str(row.get("ΒΑΡΟΣ_2",  "")).strip(),
        "ΚΡΑΜΑ_3": str(row.get("ΚΡΑΜΑ_3", "")).strip(),
        "ΒΑΡΟΣ_3":  str(row.get("ΒΑΡΟΣ_3",  "")).strip(),
    }

    print(f"\n{'='*60}")
    print(f"  MRN : {mrn}  |  PROT: {prot}")
    print(f"  PDF : {pdf1} -> {pdf}")
    print(f"{'='*60}")

    SAVE_FOLDER.mkdir(parents=True, exist_ok=True)
    if XML_PATH.exists(): XML_PATH.unlink()

    # B.1 Download XML
    try:
        driver = phase_b_download_xml(mrn, driver)
    except Exception as e:
        print(f"  Download XML failed: {e}"); traceback.print_exc(); return False

    # B.2 Parse XML
    try:
        df_result = phase_b_parse_xml()
    except Exception as e:
        print(f"  Parse XML failed: {e}"); traceback.print_exc(); return False

    # Filters
    tel_kath_val  = str(df_result["ΤΕΛ_ΚΑΘ"].iloc[0])
    dasmos_kl_val = str(df_result["ΔΑΣΜ_ΚΛ"].iloc[0])
    if tel_kath_val.startswith("71"):
        print(f"  ΤΕΛ_ΚΑΘ={tel_kath_val} — skipped (transit regime)"); return True
    if not any(dasmos_kl_val.startswith(p) for p in ALLOWED_DASMOS):
        print(f"  ΔΑΣΜ_ΚΛ={dasmos_kl_val} — outside allowed commodity codes"); return True

    # ── Alloy logic for customs office 0832 ─────────────────────────────────────
    needs_kramma = any(is_0832_kramma(r) for _, r in df_result.iterrows())

    if needs_kramma:
        num_eidi = len(df_result)

        if num_eidi == 1:
            # 1 item → use alloy data from approval popup
            aa = str(df_result.iloc[0]["Α/Α"])
            krammata = {aa: krammata_from_popup}
        else:
            # 2+ items → show secondary popup
            print(f"  Office 0832 with {num_eidi} items → ItemsPopup...")
            items_result = ItemsPopup(df_result, mrn).run()
            if items_result is None:
                print("  ItemsPopup cancelled."); return False
            krammata = items_result

        # Expand: split each item into 1-3 rows based on alloy count
        df_result = expand_krammata(df_result, krammata)
        print(f"\n  Είδη μετά expand: {len(df_result)}")

    print(f"\n{df_result[['Α/Α','ΔΑΣΜ_ΚΛ','ΒΑΡΟΣ','ΚΡΑΜΑ','ΤΙΜΗ','ΦΠΑ']].to_string(index=False)}\n")

    # B.3 PDF rename
    typos   = str(df_result.iloc[0]["ΤΥΠΟΣ"])
    dst_pdf = SAP_GUI_DIR / f"{pdf}.pdf"
    if pdf and pdf1:
        src_pdf = SAP_GUI_DIR / f"{pdf1}.pdf"
        if dst_pdf.exists():
            print(f"  PDF already renamed")
        elif src_pdf.exists():
            src_pdf.rename(dst_pdf); print(f"  PDF renamed: {pdf1} -> {pdf}")
        else:
            print(f"  PDF not found: {src_pdf}"); return False

    # B.4 SAP entry loop — one iteration per declaration line
    last_ea = ""
    try:
        for idx, (_, row_item) in enumerate(df_result.iterrows()):
            print(f"\n  Καταχώρηση {idx+1}/{len(df_result)} | Α/Α={row_item['Α/Α']} | ΚΡΑΜΑ={row_item['ΚΡΑΜΑ']} | ΒΑΡΟΣ={row_item['ΒΑΡΟΣ']}")
            if idx > 0 and not str(row_item.get("ΕΝΤ_ΑΓ", "")).strip():
                df_result.at[row_item.name, "ΕΝΤ_ΑΓ"] = last_ea
            sap_entry(df_result.loc[row_item.name], prot, sap_running=False)
            last_ea = str(df_result.at[row_item.name, "ΕΝΤ_ΑΓ"]).strip()
            sap_attach(mrn, typos)

            # Wait for successful attachment confirmation
            print("  Waiting for save...")
            while True:
                try:
                    app = Application(backend="uia").connect(
                        class_name="SAP_FRONTEND_SESSION", timeout=5)
                    app.window(class_name="SAP_FRONTEND_SESSION") \
                        .child_window(auto_id="59393") \
                        .wait("exists", timeout=5)
                    print("  Attachment OK")
                    break
                except:
                    time.sleep(1)

            # Close ZGOS attachment window
            while True:
                try:
                    get_sap_win().child_window(
                        class_name="DialogBox Container Class",
                        title_re="ZGOS_ZIMP1.*"
                    ).wait("exists", timeout=5).close()
                    break
                except:
                    break

            time.sleep(2)

    except Exception as e:
        print(f"  SAP entry failed: {e}"); traceback.print_exc()
        close_sap(); return False

    close_sap(); time.sleep(2)

    # B.5 Archive PDF
    kath        = int(df_result.iloc[0]["ΚΑΘ"])
    dest_subdir = KATH_DIR.get(kath)
    if dest_subdir and dst_pdf.exists():
        dest = ATLAS_BASE / dest_subdir / f"{pdf}.pdf"
        try:
            shutil.move(str(dst_pdf), str(dest))
            print(f"  PDF moved -> {dest_subdir}")
        except Exception as e:
            print(f"  Move failed: {e}")
    elif not dest_subdir:
        print(f"  Unknown customs regime KAΘ={kath}")

    return True

# ==============================================================================
# MAIN
# ==============================================================================

def ask_start_mode() -> int:
    root = tk.Tk()
    root.title("Εκκίνηση")
    root.resizable(False, False)
    root.configure(bg="#F5F4F0")
    root.attributes("-topmost", True)
    w, h = 520, 220
    sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    choice = tk.IntVar(value=0)

    tk.Label(root, text="Από πού να ξεκινήσω;",
        bg="#F5F4F0", fg="#1A1A1A", font=("Consolas", 11, "bold")
    ).pack(pady=(20, 12))

    for val, label, bg, fg in [
        (1, "1 — Πλήρης ροή  (ICISnet + SAP Export + Queries + Popup)", "#1A1A1A", "#FFFFFF"),
        (2, "2 — Queries από αποθηκευμένο PDF + Database", "#2A2A2A", "#FFFFFF"),
        (3, "3 — Από Popup  (έχω ήδη FULL_RESULTS)", "#E8E6DF", "#1A1A1A"),
    ]:
        tk.Button(root, text=label,
            font=("Consolas", 10), bg=bg, fg=fg,
            relief="flat", padx=12, pady=7, cursor="hand2",
            command=lambda v=val: [choice.set(v), root.destroy()]
        ).pack(fill="x", padx=30, pady=(0, 5))

    root.lift(); root.focus_force()
    root.mainloop()
    return choice.get()


def read_icisnet_from_pdf() -> pd.DataFrame:
    import pdfplumber
    rows = []
    with pdfplumber.open(PDF_SAVE_PATH) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split("\n"):
                parts = line.split()
                if len(parts) < 3:
                    continue
                if parts[0] in ("LRN", "Αποτελέσματα", "MRN"):
                    continue
                lrn = parts[0]
                mrn = parts[1] if len(parts) > 1 and parts[1].startswith("26GR") else ""
                typos = next((x for x in parts if x.startswith("IM-")), "")
                if not typos:
                    continue
                rows.append({"LRN": lrn, "MRN": mrn, "Τύπος Δήλωσης": typos})

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["MRN","ΤΥΠΟΣ","ΚΑΤΑΣΤΑΣΗ","LRN","ΗΜ_ΥΠΟΒ","ΗΜ_ΕΝΗΜ","PDF"])

    repl = [
        ("YOUR_LRN_PREFIX_ALT/", "YOUR_LRN_PREFIX/"), ("ELVELV", "ELV"),
        ("YOUR_VAT_PREFIX_WRONG", "YOUR_VAT_PREFIX_CORRECT"),
    ]
    df["LRN"] = df["LRN"].astype(str)
    for old, new in repl:
        df["LRN"] = df["LRN"].str.replace(old, new, regex=False)

    df["Τύπος Δήλωσης"] = df["Τύπος Δήλωσης"].str.replace("-", "", regex=False)
    df["ΚΑΤΑΣΤΑΣΗ"] = "ID29"
    df["ΗΜ_ΥΠΟΒ"] = pd.NaT
    df["ΗΜ_ΕΝΗΜ"] = pd.NaT
    df["PDF"] = df["MRN"].astype(str) + " " + df["Τύπος Δήλωσης"].astype(str)
    df = df.rename(columns={"Τύπος Δήλωσης": "ΤΥΠΟΣ"})
    df = df[~df["LRN"].str.contains("EXCLUDED_LRN_PREFIX", na=False)]
    df = df[df["LRN"].str.contains(r"ELV|ΕLV", na=False)]
    df = df[df["MRN"] != ""]

    return df[["MRN","ΤΥΠΟΣ","ΚΑΤΑΣΤΑΣΗ","LRN","ΗΜ_ΥΠΟΒ","ΗΜ_ΕΝΗΜ","PDF"]].reset_index(drop=True)


def main():
    T0 = perf_counter()
    setup_keyboard()

    mode = ask_start_mode()
    if mode == 0:
        return

    if mode == 1:
        print("\n── PHASE A: IMPORT DECLARATIONS ──")
        print("\n[1/3] ICISnet scraping...")
        df_final = phase_a_icisnet()
        print("\n[2/3] SAP Export...")
        phase_a_sap_export()
        print("\n[3/3] Queries -> FULL_RESULTS.xlsx...")
        df_opened = phase_a_queries(df_final)

    elif mode == 2:
        print("\n[Queries] from saved PDF...")
        df_final = read_icisnet_from_pdf()
        print(f"  PDF: {len(df_final)} records")
        df_opened = phase_a_queries(df_final)

    else:
        print("\n── Loading FULL_RESULTS.xlsx ──")
        df_opened = pd.read_excel(OUTPUT_EXCEL, sheet_name="opened")
        df_opened.columns = df_opened.columns.str.strip()

    if df_opened.empty:
        show_info("⚠️  No declarations pending for entry.")
        return


    # ── POPUP ─────────────────────────────────────────────────────────────────
    print("\n── POPUP: Waiting for user ──")

    rows = []
    for _, r in df_opened.iterrows():
        d = {}
        for col in SHOW_COLS:
            val = r.get(col, "")
            if pd.isna(val): val = ""
            if col == "PROT" and str(val).endswith(".0"):
                val = str(int(float(val)))
            d[col] = str(val).strip() if str(val).strip() != "nan" else ""
        for extra in df_opened.columns:
            if extra not in d:
                v = r.get(extra, "")
                d[extra] = "" if pd.isna(v) else v
        rows.append(d)
  
    selected = ApprovalPopup(rows).run()

    if selected is None:
        print("  Cancelled.")
        return

    print(f"  {len(selected)} MRNs selected — starting Phase B...")

    # Άνοιγμα browser και login μία φορά
    print("  Opening browser and logging into ICISnet...")
    driver = make_chrome(download_folder=SAVE_FOLDER)
    wait = WebDriverWait(driver, 45)
    driver.get("https://www1.gsis.gr/icisnet/itrader/common/home.jsf")
    wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(ICISNET_USER)
    driver.find_element(By.NAME, "password").send_keys(ICISNET_PASS)
    driver.find_element(By.NAME, "btn_login").click()
    time.sleep(3)
    print("  ICISnet login OK!")

    # ── ΦΑΣΗ Β ───────────────────────────────────────────────────────────────
    print("\n── PHASE B: SAP Entries ──")

    df_excel = pd.read_excel(OUTPUT_EXCEL, sheet_name="opened")
    if "Status" not in df_excel.columns:
        df_excel["Status"] = ""

    ok_n = 0; fail_n = 0

    for i, row in enumerate(selected, 1):
        mrn = row.get("MRN", "???")
        print(f"\n[{i}/{len(selected)}] {mrn}")

        success = process_mrn(row, driver)

        if success:
            ok_n += 1
            try:
                df_excel.loc[df_excel["MRN"].astype(str) == str(mrn), "Status"] = "DONE"
                with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl",
                                    mode="a", if_sheet_exists="replace") as w:
                    df_excel.to_excel(w, sheet_name="opened", index=False)
                print(f"  Status -> DONE")
            except Exception as e:
                print(f"  Status update failed: {e}")
        else:
            fail_n += 1
            print(f"  FAIL — continuing to next...")

    try:
        driver.quit()
        print("  Browser closed")
    except:
        pass

    elapsed = perf_counter() - T0
    print(f"\n{'='*60}")
    print(f"  ΟΛΟΚΛΗΡΩΘΗΚΕ  |  OK:{ok_n}  FAIL:{fail_n}  |  {elapsed/60:.1f} λεπτά")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
